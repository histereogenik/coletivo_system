from __future__ import annotations

from decimal import Decimal
from io import BytesIO
from typing import Iterable

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


def cents_to_reais(value_cents: int | None) -> Decimal | str:
    if value_cents is None:
        return ""
    return Decimal(value_cents) / Decimal(100)


def create_xlsx_response(
    filename: str, headers: list[str], rows: Iterable[Iterable]
) -> HttpResponse:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(headers)

    for cell in worksheet[1]:
        cell.font = Font(bold=True)

    for row in rows:
        worksheet.append(list(row))

    for idx, column in enumerate(worksheet.columns, start=1):
        max_len = 0
        for cell in column:
            if cell.value is None:
                continue
            max_len = max(max_len, len(str(cell.value)))
        worksheet.column_dimensions[get_column_letter(idx)].width = min(max_len + 2, 50)

    buffer = BytesIO()
    workbook.save(buffer)
    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}.xlsx"'
    return response
